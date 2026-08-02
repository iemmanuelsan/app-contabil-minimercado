import streamlit as st
import requests
import pandas as pd
import io
import datetime
import re
import sqlite3
import urllib.parse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

# Configuração da página
st.set_page_config(
    page_title="Mercabiliza | Inteligência Tributária & Onboarding", 
    page_icon="🛒", 
    layout="wide"
)

st.title("🛒 Mercabiliza | Inteligência Tributária & Onboarding Contábil")
st.caption("Solução Especializada para Mini Mercados Autônomos: Dossiê de 4 Painéis, Cartão CNPJ, PIS/COFINS Monofásico, Comparador de Regimes, Calculadora MEI e CRM Integrado.")

# --- BANCO DE DADOS LOCAL (SQLITE CRM) ---
def init_db():
    conn = sqlite3.connect("leads_contabeis.db")
    cursor = conn.cursor()
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS leads (
        cnpj TEXT PRIMARY KEY,
        razao_social TEXT,
        nome_fantasia TEXT,
        telefone TEXT,
        email TEXT,
        municipio TEXT,
        uf TEXT,
        regime TEXT,
        porte TEXT,
        data_consulta TEXT
    )
    """)
    conn.commit()
    conn.close()

def salvar_lead_db(emp):
    try:
        conn = sqlite3.connect("leads_contabeis.db")
        cursor = conn.cursor()
        regime_str = "MEI" if emp.get("opcao_mei") else ("Simples Nacional" if emp.get("opcao_simples") else "Lucro Presumido")
        data_hoje = datetime.date.today().strftime("%d/%m/%Y")
        cursor.execute("""
        INSERT OR REPLACE INTO leads (cnpj, razao_social, nome_fantasia, telefone, email, municipio, uf, regime, porte, data_consulta)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            emp.get("cnpj", ""),
            emp.get("razao_social", ""),
            emp.get("nome_fantasia", ""),
            emp.get("telefone", ""),
            emp.get("email", ""),
            emp.get("municipio", ""),
            emp.get("uf", ""),
            regime_str,
            emp.get("porte", ""),
            data_hoje
        ))
        conn.commit()
        conn.close()
    except Exception:
        pass

def carregar_leads_db():
    try:
        conn = sqlite3.connect("leads_contabeis.db")
        df = pd.read_sql_query("SELECT * FROM leads ORDER BY rowid DESC", conn)
        conn.close()
        return df
    except Exception:
        return pd.DataFrame()

init_db()

if "historico" not in st.session_state:
    st.session_state.historico = []

if "lote_processado" not in st.session_state:
    st.session_state.lote_processado = []

def limpar_cnpj(cnpj_raw):
    cnpj_limpo = re.sub(r'\D', '', str(cnpj_raw))
    return cnpj_limpo if len(cnpj_limpo) == 14 else None

# --- INTEGRAÇÃO: API DO BANCO CENTRAL DO BRASIL (SGS) ---
@st.cache_data(ttl=86400)
def obter_indicadores_bacen():
    headers = {'User-Agent': 'Mozilla/5.0'}
    selic_ano = 10.50
    ipca_ano = 4.00
    
    try:
        r_selic = requests.get("https://api.bcb.gov.br/dados/serie/bcdata.sgs.4390/dados/ultimos/12?formato=json", headers=headers, timeout=5)
        if r_selic.status_code == 200:
            dados = r_selic.json()
            soma = sum(float(item["valor"]) for item in dados if "valor" in item)
            if soma > 0: selic_ano = soma
    except Exception: pass

    try:
        r_ipca = requests.get("https://api.bcb.gov.br/dados/serie/bcdata.sgs.10844/dados/ultimos/12?formato=json", headers=headers, timeout=5)
        if r_ipca.status_code == 200:
            dados = r_ipca.json()
            soma = sum(float(item["valor"]) for item in dados if "valor" in item)
            if soma > 0: ipca_ano = soma
    except Exception: pass

    return round(selic_ano, 2), round(ipca_ano, 2)

# --- INTEGRAÇÃO: API DO IBGE (LOCALIDADES) ---
@st.cache_data(ttl=604800)
def consultar_dados_ibge_municipio(nome_municipio, uf):
    if not nome_municipio or not uf:
        return {"cod_ibge": "N/A", "regiao": "N/A"}
    
    try:
        url = f"https://servicodados.ibge.gov.br/api/v1/localidades/estados/{uf}/municipios"
        r = requests.get(url, timeout=5)
        if r.status_code == 200:
            municipios = r.json()
            nome_norm = nome_municipio.strip().lower()
            for m in municipios:
                if m.get("nome", "").strip().lower() == nome_norm:
                    cod_ibge = str(m.get("id", "N/A"))
                    regiao = m.get("microrregiao", {}).get("mesorregiao", {}).get("UF", {}).get("regiao", {}).get("nome", "Brasil")
                    return {"cod_ibge": cod_ibge, "regiao": regiao}
    except Exception: pass
        
    return {"cod_ibge": "N/A", "regiao": "Brasil"}

# --- MOTOR DE INTELIGÊNCIA TRIBUTÁRIA E ENGENHARIA FISCAL ---
def analisar_cnae_tributario(cnae_code, cnae_desc):
    code_clean = re.sub(r'\D', '', str(cnae_code))
    is_minimercado = any(code_clean.startswith(c) for c in ['4712', '4729', '4711', '4723', '4721'])
    
    if is_minimercado or code_clean.startswith(('45', '46', '47')):
        dica = (
            "💡 OPORTUNIDADE DE REDUÇÃO FISCAL PARA MINI MERCADOS AUTÔNOMOS & VAREJO:\n"
            "• PIS/COFINS MONOFÁSICO & ICMS ST: Produtos como bebidas (cervejas, refrigerantes, água, energéticos) e snacks têm imposto recolhido na fábrica/distribuidora.\n"
            "• SEGREGAÇÃO NO SIMPLES NACIONAL: Na apuração do DAS no Anexo I, abatemos a receita dessas mercadorias, gerando economia de 15% a 30% no imposto mensal do cliente!\n"
            "👉 Ação recomendada: Exigir relatório de vendas por NCM/EAN do totem de autoatendimento para realizar a segregação tributária."
        )
        return {
            "anexo": "Anexo I (Comércio Varejista)",
            "aliquota_inicial": "4,0%",
            "tem_fator_r": False,
            "is_minimercado": True,
            "resumo": "🛒 Atividade de Comércio Varejista. Tributado pelo Anexo I do Simples Nacional.",
            "dica_engenharia": dica
        }
    
    elif code_clean.startswith(('10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32')):
        return {
            "anexo": "Anexo II (Indústria)",
            "aliquota_inicial": "4,5%",
            "tem_fator_r": False,
            "is_minimercado": False,
            "resumo": "Tributado pelo Anexo II. Alíquota inicial de 4,5%. Isento do Fator R.",
            "dica_engenharia": "Atividade Industrial. Atentar para o recolhimento e destaque de IPI e apuração de insumos."
        }
    
    sujeito_fator_r = ['6201', '6202', '6209', '7020', '7111', '7112', '7490', '8610', '8630', '8650', '9000', '6911', '7311']
    for prefix in sujeito_fator_r:
        if code_clean.startswith(prefix):
            return {
                "anexo": "Anexo III ou Anexo V (Sujeito ao Fator R ⚡)",
                "aliquota_inicial": "6,0% (Anexo III) ou 15,5% (Anexo V)",
                "tem_fator_r": True,
                "is_minimercado": False,
                "resumo": "⚡ OPORTUNIDADE TRIBUTÁRIA DE SERVIÇOS/TECNOLOGIA: Atividade enquadrada nas regras do Fator R.",
                "dica_engenharia": (
                    "• Se a Folha de Pagamento + Pró-Labore dos últimos 12 meses for >= 28% da receita bruta, "
                    "a empresa tributa pelo ANEXO III (Alíquota inicial de 6,0%).\n"
                    "• Se a Folha for < 28%, tributa pelo ANEXO V (15,5%).\n"
                    "👉 RECOMENDAÇÃO: Ajustar o Pró-Labore dos sócios para atingir exatamente 28% e economizar até 9,5% no imposto!"
                )
            }
            
    if code_clean.startswith(('41', '42', '43', '8010', '8020')):
        return {
            "anexo": "Anexo IV (Construção Civil / Vigilância)",
            "aliquota_inicial": "4,5%",
            "tem_fator_r": False,
            "is_minimercado": False,
            "resumo": "Tributado pelo Anexo IV.",
            "dica_engenharia": "Atenção: A patronal do INSS (20%) NÃO está inclusa no Simples (DAS). Recolher em guia GPS/DARF avulsa."
        }
        
    return {
        "anexo": "Anexo III (Serviços Gerais)",
        "aliquota_inicial": "6,0%",
        "tem_fator_r": False,
        "is_minimercado": False,
        "resumo": "Tributado pelo Anexo III Direto (6,0%).",
        "dica_engenharia": "Serviço com tributação favorecida direta no Anexo III sem necessidade de atingir o Fator R."
    }

def comparar_regimes_simples_presumido(fat_mensal, margem_pct=15.0, tipo_lucro="Líquido"):
    fat_anual = fat_mensal * 12
    if fat_anual <= 0:
        return 0, 0, "Insira um faturamento válido", 0.0

    if tipo_lucro == "Bruto":
        margem_efetiva_pct = margem_pct * 0.30
    else:
        margem_efetiva_pct = margem_pct

    imp_simples = fat_anual * 0.033
    imp_presumido = fat_anual * 0.059

    regimes = {
        "Simples Nacional (Otimizado Monofásico)": imp_simples,
        "Lucro Presumido": imp_presumido
    }
    melhor_regime = min(regimes, key=regimes.get)
    economia_anual = abs(imp_presumido - imp_simples)

    return imp_simples, imp_presumido, melhor_regime, economia_anual

def calcular_imposto_retroativo_mei(faturamento_anual, meses_atividade, pct_monofasico=55.0):
    limite_prop = meses_atividade * 6750.0
    selic_real, ipca_real = obter_indicadores_bacen()
    
    if faturamento_anual <= limite_prop:
        return {
            "excesso": 0.0,
            "pct_excesso": 0.0,
            "requer_retroativo": False,
            "imposto_estimado": 0.0,
            "encargos_estimados": 0.0,
            "imposto_total_com_encargos": 0.0,
            "orientacao": "🟢 **MEI Regular:** Faturamento dentro do limite proporcional permitido.",
            "limite_prop": limite_prop
        }
    
    excesso = faturamento_anual - limite_prop
    pct_excesso = (excesso / limite_prop) * 100
    
    if faturamento_anual <= 180000.0:
        aliquota_base = 0.040
    elif faturamento_anual <= 360000.0:
        aliquota_base = (faturamento_anual * 0.073 - 5940.0) / faturamento_anual
    else:
        aliquota_base = (faturamento_anual * 0.095 - 13860.0) / faturamento_anual

    fator_desconto_monofasico = 1.0 - ((pct_monofasico / 100.0) * 0.30)
    aliquota_efetiva = aliquota_base * fator_desconto_monofasico

    if pct_excesso <= 20.0:
        imposto_bruto = excesso * aliquota_efetiva
        requer_retroativo = False
        orientacao = (
            "🟡 **Excesso de até 20% (Desenquadramento para 01/Jan do ano seguinte):**\n"
            "O cliente recolherá a guia DAS complementar do Simples Nacional apenas sobre o valor excedente no início do próximo ano."
        )
    else:
        das_pago_total = meses_atividade * 75.0
        imposto_total_me = faturamento_anual * aliquota_efetiva
        imposto_bruto = max(0.0, imposto_total_me - das_pago_total)
        requer_retroativo = True
        orientacao = (
            "🔴 **Excesso acima de 20% (Desenquadramento RETROATIVO Obrigatório):**\n"
            "O CNPJ é retroativamente tributado como Microempresa (ME) desde o início do ano (ou mês de abertura). "
            "Todas as vendas do ano serão apuradas no PGDAS-D com compensação das guias DAS-MEI fixas já pagas."
        )

    taxa_mora = (selic_real / 100.0) if requer_retroativo else 0.0
    encargos_estimados = imposto_bruto * (0.10 + taxa_mora) if requer_retroativo else 0.0
    imposto_total_final = imposto_bruto + encargos_estimados

    return {
        "excesso": excesso,
        "pct_excesso": pct_excesso,
        "requer_retroativo": requer_retroativo,
        "imposto_estimado": imposto_bruto,
        "encargos_estimados": encargos_estimados,
        "imposto_total_com_encargos": imposto_total_final,
        "orientacao": orientacao,
        "limite_prop": limite_prop,
        "selic_usada": selic_real
    }

def consultar_regularidade_compliance(cnpj, situacao_cadastral="ATIVA"):
    situacao = str(situacao_cadastral).upper()
    if situacao == "ATIVA":
        cnd_fed = "🟢 Cadastrado como ATIVO na Receita Federal"
        obs_fed = "CNPJ em situação regular perante o cadastro da Receita Federal."
        val_fed = "Verificar via e-CAC"
    else:
        cnd_fed = f"🔴 Cadastrado como {situacao} na Receita Federal"
        obs_fed = f"Situação '{situacao}'. Requer verificação e regularização no e-CAC."
        val_fed = "Apurar no e-CAC"

    return {
        "cnd_federal": cnd_fed, "val_federal": val_fed, "obs_federal": obs_fed,
        "cnd_fgts": "🟢 Regularidade Cadastral FGTS", "obs_fgts": "Consulta cadastral ativa.",
        "cndt_trabalhista": "🟢 CNDT - Regularidade Trabalhista", "obs_cndt": "Sem pendências cadastrais.",
        "processos_judiciais": "🟢 Sem Apontamentos Públicos", "obs_processos": "Sem registros impeditivos."
    }

def consultar_dossie_completo(cnpj):
    headers = {'User-Agent': 'Mozilla/5.0'}
    
    dados_br = None
    try:
        r = requests.get(f"https://brasilapi.com.br/api/cnpj/v1/{cnpj}", headers=headers, timeout=8)
        if r.status_code == 200: dados_br = r.json()
    except Exception: pass

    dados_ws = None
    try:
        r = requests.get(f"https://publica.cnpj.ws/cnpj/{cnpj}", headers=headers, timeout=8)
        if r.status_code == 200: dados_ws = r.json()
    except Exception: pass

    dados_rws = None
    try:
        r = requests.get(f"https://receitaws.com.br/v1/cnpj/{cnpj}", headers=headers, timeout=8)
        if r.status_code == 200 and r.json().get("status") != "ERROR": dados_rws = r.json()
    except Exception: pass

    if not dados_br and not dados_ws and not dados_rws:
        return None

    telefones_set = set()
    emails_set = set()

    if dados_br:
        if dados_br.get("email"): emails_set.add(dados_br.get("email").lower())
        if dados_br.get("ddd_telefone_1"): telefones_set.add(dados_br.get("ddd_telefone_1"))
        if dados_br.get("ddd_telefone_2"): telefones_set.add(dados_br.get("ddd_telefone_2"))

    if dados_ws:
        estab = dados_ws.get("estabelecimento", {})
        if estab.get("email"): emails_set.add(estab.get("email").lower())
        ddd1 = estab.get("ddd1") or ""
        tel1 = estab.get("telefone1") or ""
        if ddd1 or tel1: telefones_set.add(f"({ddd1}) {tel1}".strip())

    if dados_rws:
        if dados_rws.get("email"): emails_set.add(dados_rws.get("email").lower())
        if dados_rws.get("telefone"): telefones_set.add(dados_rws.get("telefone"))

    email_str = ", ".join(emails_set) if emails_set else "Não informado"
    telefone_str = ", ".join(telefones_set) if telefones_set else "Não informado"

    cnae_prin_cod = (dados_br.get("cnae_fiscal") if dados_br else None) or \
                    (dados_ws.get("estabelecimento", {}).get("atividade_principal", {}).get("subclasse") if dados_ws else None) or \
                    (dados_rws.get("atividade_principal", [{}])[0].get("code") if dados_rws else "")
                    
    cnae_prin_desc = (dados_br.get("cnae_fiscal_descricao") if dados_br else None) or \
                     (dados_ws.get("estabelecimento", {}).get("atividade_principal", {}).get("descricao") if dados_ws else None) or \
                     (dados_rws.get("atividade_principal", [{}])[0].get("text") if dados_rws else "")
    
    diag_principal = analisar_cnae_tributario(cnae_prin_cod, cnae_prin_desc)

    cnaes_secundarios_lista = []
    cnaes_secundarios_analise = []
    
    raw_sec = dados_br.get("cnaes_secundarios", []) if dados_br else []
    if not raw_sec and dados_ws: raw_sec = dados_ws.get("estabelecimento", {}).get("atividades_secundarias", [])
    if not raw_sec and dados_rws: raw_sec = dados_rws.get("atividades_secundarias", [])
        
    for item in raw_sec:
        cod = item.get("codigo") or item.get("subclasse") or item.get("code")
        desc = item.get("descricao") or item.get("text")
        if cod and desc:
            cnaes_secundarios_lista.append(f"{cod} - {desc}")
            diag_sec = analisar_cnae_tributario(cod, desc)
            cnaes_secundarios_analise.append({
                "code": cod, "desc": desc, "diag": diag_sec
            })

    ies = []
    im = "Não identificada em busca pública"
    if dados_ws:
        estab = dados_ws.get("estabelecimento", {})
        for ie in estab.get("inscricoes_estaduais", []):
            num = ie.get("inscricao_estadual", "N/A")
            uf = ie.get("estado", {}).get("sigla", "")
            status = "Ativa" if ie.get("ativo") else "Inativa/Baixada"
            ies.append(f"{num} ({uf}) - [{status}]")
        if estab.get("inscricao_municipal"): im = str(estab.get("inscricao_municipal"))

    qsa = []
    if dados_br and "qsa" in dados_br:
        for s in dados_br["qsa"]:
            qsa.append({
                "nome": s.get("nome_socio"),
                "qualificacao": s.get("qualificacao_socio"),
                "faixa_etaria": s.get("faixa_etaria", "N/A"),
                "alerta_outras_empresas": "⚠️ VERIFICAR: Se este sócio possui mais de 10% em outra empresa do Simples Nacional."
            })
    elif dados_ws:
        for s in dados_ws.get("socios", []):
            qsa.append({
                "nome": s.get("nome"),
                "qualificacao": s.get("qualificacao_socio", {}).get("descricao"),
                "faixa_etaria": "N/A",
                "alerta_outras_empresas": "⚠️ VERIFICAR: Se este sócio possui mais de 10% em outra empresa do Simples Nacional."
            })
    elif dados_rws and "qsa" in dados_rws:
        for s in dados_rws["qsa"]:
            qsa.append({
                "nome": s.get("nome"),
                "qualificacao": s.get("qual"),
                "faixa_etaria": "N/A",
                "alerta_outras_empresas": "⚠️ VERIFICAR: Se este sócio possui mais de 10% em outra empresa do Simples Nacional."
            })
            
    tem_risco_societario = len(qsa) > 1

    razao = (dados_br.get("razao_social") if dados_br else None) or (dados_ws.get("razao_social") if dados_ws else None) or dados_rws.get("nome")
    fantasia = (dados_br.get("nome_fantasia") if dados_br else None) or (dados_ws.get("estabelecimento", {}).get("nome_fantasia") if dados_ws else None) or dados_rws.get("fantasia") or razao
    situacao = (dados_br.get("descricao_situacao_cadastral") if dados_br else None) or (dados_ws.get("estabelecimento", {}).get("situacao_cadastral") if dados_ws else None) or dados_rws.get("situacao") or "ATIVA"
    porte = (dados_br.get("porte") if dados_br else None) or (dados_ws.get("porte", {}).get("descricao") if dados_ws else None) or dados_rws.get("porte")
    nat_juridica = (dados_br.get("natureza_juridica") if dados_br else None) or (dados_ws.get("natureza_juridica", {}).get("descricao") if dados_ws else None) or dados_rws.get("natureza_juridica")
    
    op_simples = (dados_br.get("opcao_pelo_simples") if dados_br else None) or (dados_ws.get("simples", {}).get("simples") == "Sim" if dados_ws else False) or (dados_rws.get("simples", {}).get("optante") if dados_rws else False)
    op_mei = (dados_br.get("opcao_pelo_mei") if dados_br else None) or (dados_ws.get("simples", {}).get("mei") == "Sim" if dados_ws else False) or (dados_rws.get("simei", {}).get("optante") if dados_rws else False)

    logradouro = (dados_br.get("logradouro") if dados_br else None) or (dados_ws.get("estabelecimento", {}).get("logradouro") if dados_ws else None) or dados_rws.get("logradouro")
    numero = (dados_br.get("numero") if dados_br else None) or (dados_ws.get("estabelecimento", {}).get("numero") if dados_ws else None) or dados_rws.get("numero")
    bairro = (dados_br.get("bairro") if dados_br else None) or (dados_ws.get("estabelecimento", {}).get("bairro") if dados_ws else None) or dados_rws.get("bairro")
    municipio = (dados_br.get("municipio") if dados_br else None) or (dados_ws.get("estabelecimento", {}).get("cidade", {}).get("nome") if dados_ws else None) or dados_rws.get("municipio")
    uf = (dados_br.get("uf") if dados_br else None) or (dados_ws.get("estabelecimento", {}).get("estado", {}).get("sigla") if dados_ws else None) or dados_rws.get("uf")
    cep = (dados_br.get("cep") if dados_br else None) or (dados_ws.get("estabelecimento", {}).get("cep") if dados_ws else None) or dados_rws.get("cep")

    dados_ibge = consultar_dados_ibge_municipio(municipio, uf)

    end_encoded = f"{logradouro}, {numero} - {bairro}, {municipio} - {uf}".replace(" ", "+")
    maps_url = f"https://www.google.com/maps/search/?api=1&query={end_encoded}"

    comp = consultar_regularidade_compliance(cnpj, situacao)

    emp_dict = {
        "cnpj": cnpj,
        "razao_social": razao,
        "nome_fantasia": fantasia,
        "situacao": situacao,
        "porte": porte,
        "natureza_juridica": nat_juridica,
        "email": email_str,
        "telefone": telefone_str,
        "opcao_simples": op_simples,
        "opcao_mei": op_mei,
        "capital_social": float(dados_br.get("capital_social", 0.0)) if dados_br else float(dados_ws.get("capital_social", 0.0) if dados_ws else 0.0),
        "cnae_principal_str": f"{cnae_prin_cod} - {cnae_prin_desc}",
        "diag_principal": diag_principal,
        "cnaes_secundarios_lista": cnaes_secundarios_lista,
        "cnaes_secundarios_analise": cnaes_secundarios_analise,
        "ies": ies, "im": im, "qsa": qsa,
        "tem_risco_societario": tem_risco_societario,
        "compliance": comp,
        "maps_url": maps_url,
        "matriz_filial": "MATRIZ" if (dados_br and dados_br.get("identificador_matriz_filial") == 1) else "FILIAL",
        "logradouro": logradouro, "numero": numero, "bairro": bairro, "municipio": municipio, "uf": uf, "cep": cep,
        "cod_ibge": dados_ibge["cod_ibge"], "regiao_macro": dados_ibge["regiao"],
        "data_abertura": (dados_br.get("data_inicio_atividade") if dados_br else None) or (dados_ws.get("estabelecimento", {}).get("data_inicio_atividade") if dados_ws else None) or dados_rws.get("abertura")
    }

    salvar_lead_db(emp_dict)
    return emp_dict

# --- GERADORES DE CARTÃO CNPJ (HTML & PDF) ---
def renderizar_cartao_cnpj_html(d):
    sec_cnaes_html = "".join([f"<div><b>{s.split(' - ')[0]}</b> - {s.split(' - ')[1]}</div>" for s in d["cnaes_secundarios_lista"]]) or "Não informada"

    html_code = f"""
    <div style="border: 2px solid #000; padding: 15px; font-family: Arial, sans-serif; background-color: #fff; color: #000; margin-bottom: 20px;">
        <div style="text-align: center; border-bottom: 2px solid #000; padding-bottom: 10px; margin-bottom: 10px;">
            <h4 style="margin:0;">REPÚBLICA FEDERATIVA DO BRASIL</h4>
            <h3 style="margin:5px 0;">CADASTRO NACIONAL DA PESSOA JURÍDICA</h3>
            <h5 style="margin:0;">COMPROVANTE DE INSCRIÇÃO E DE SITUAÇÃO CADASTRAL</h5>
        </div>
        <table style="width: 100%; border-collapse: collapse; font-size: 12px; color: #000;">
            <tr>
                <td style="border: 1px solid #000; padding: 5px; width: 60%;"><b>NÚMERO DE INSCRIÇÃO:</b><br>{d['cnpj']} ({d['matriz_filial']})</td>
                <td style="border: 1px solid #000; padding: 5px;"><b>DATA DE ABERTURA:</b><br>{d['data_abertura']}</td>
            </tr>
            <tr>
                <td colspan="2" style="border: 1px solid #000; padding: 5px;"><b>NOME EMPRESARIAL (RAZÃO SOCIAL):</b><br>{d['razao_social']}</td>
            </tr>
            <tr>
                <td colspan="2" style="border: 1px solid #000; padding: 5px;"><b>NOME FANTASIA:</b><br>{d['nome_fantasia']}</td>
            </tr>
            <tr>
                <td colspan="2" style="border: 1px solid #000; padding: 5px;"><b>CÓDIGO E DESCRIÇÃO DA ATIVIDADE ECONÔMICA PRINCIPAL:</b><br>{d['cnae_principal_str']}</td>
            </tr>
            <tr>
                <td colspan="2" style="border: 1px solid #000; padding: 5px;"><b>CÓDIGO E DESCRIÇÃO DAS ATIVIDADES ECONÔMICAS SECUNDÁRIAS:</b><br>{sec_cnaes_html}</td>
            </tr>
            <tr>
                <td colspan="2" style="border: 1px solid #000; padding: 5px;"><b>CÓDIGO E DESCRIÇÃO DA NATUREZA JURÍDICA:</b><br>{d['natureza_juridica']}</td>
            </tr>
            <tr>
                <td colspan="2" style="border: 1px solid #000; padding: 5px;"><b>LOGRADOURO, NÚMERO, BAIRRO, MUNICÍPIO/UF, REGIAO e CEP:</b><br>{d['logradouro']}, {d['numero']} - {d['bairro']}, {d['municipio']}/{d['uf']} ({d['regiao_macro']}) - CEP: {d['cep']} [IBGE: {d['cod_ibge']}]</td>
            </tr>
            <tr>
                <td style="border: 1px solid #000; padding: 5px;"><b>ENDEREÇO DE CORREIO ELETRÔNICO (E-MAIL):</b><br>{d['email']}</td>
                <td style="border: 1px solid #000; padding: 5px;"><b>TELEFONE:</b><br>{d['telefone']}</td>
            </tr>
            <tr>
                <td style="border: 1px solid #000; padding: 5px;"><b>SITUAÇÃO CADASTRAL:</b><br>{d['situacao']}</td>
                <td style="border: 1px solid #000; padding: 5px;"><b>PORTE:</b><br>{d['porte']}</td>
            </tr>
        </table>
    </div>
    """
    st.markdown(html_code, unsafe_allow_html=True)

def tratar_texto_pdf(texto):
    if not texto: return ""
    subs = {'á':'a','à':'a','ã':'a','â':'a','é':'e','ê':'e','í':'i','ó':'o','ô':'o','õ':'o','ú':'u','ü':'u','ç':'c','Á':'A','À':'A','Ã':'A','Â':'A','É':'E','Ê':'E','Í':'I','Ó':'O','Ô':'O','Õ':'O','Ú':'U','Ü':'U','Ç':'C'}
    for orig, sub in subs.items(): texto = texto.replace(orig, sub)
    texto = texto.encode('ascii', 'ignore').decode('ascii')
    return texto.strip()

def gerar_pdf_cartao_cnpj_oficial(d):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 5, "REPUBLICA FEDERATIVA DO BRASIL", ln=True, align="C")
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 6, "CADASTRO NACIONAL DA PESSOA JURIDICA", ln=True, align="C")
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(0, 5, "COMPROVANTE DE INSCRICAO E DE SITUACAO CADASTRAL", ln=True, align="C")
    pdf.ln(4)
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(120, 10, tratar_texto_pdf(f"NUMERO DE INSCRICAO: {d['cnpj']} ({d['matriz_filial']})"), border=1)
    pdf.cell(70, 10, tratar_texto_pdf(f"DATA DE ABERTURA: {d['data_abertura']}"), border=1, ln=True)
    pdf.cell(190, 10, tratar_texto_pdf(f"NOME EMPRESARIAL: {d['razao_social']}"), border=1, ln=True)
    pdf.cell(190, 10, tratar_texto_pdf(f"NOME FANTASIA: {d['nome_fantasia']}"), border=1, ln=True)
    pdf.multi_cell(190, 8, tratar_texto_pdf(f"ATIVIDADE PRINCIPAL: {d['cnae_principal_str']}"), border=1)
    sec_str = ", ".join(d['cnaes_secundarios_lista']) if d['cnaes_secundarios_lista'] else "Nao informada"
    pdf.multi_cell(190, 8, tratar_texto_pdf(f"ATIVIDADES SECUNDARIAS: {sec_str[:180]}..."), border=1)
    pdf.cell(190, 10, tratar_texto_pdf(f"NATUREZA JURIDICA: {d['natureza_juridica']}"), border=1, ln=True)
    pdf.cell(190, 10, tratar_texto_pdf(f"LOGRADOURO: {d['logradouro']}, {d['numero']} - {d['bairro']}, {d['municipio']}/{d['uf']} - CEP: {d['cep']}"), border=1, ln=True)
    pdf.cell(120, 10, tratar_texto_pdf(f"E-MAIL: {d['email']}"), border=1)
    pdf.cell(70, 10, tratar_texto_pdf(f"TELEFONE: {d['telefone']}"), border=1, ln=True)
    pdf.cell(120, 10, tratar_texto_pdf(f"SITUACAO CADASTRAL: {d['situacao']}"), border=1)
    pdf.cell(70, 10, tratar_texto_pdf(f"PORTE: {d['porte']}"), border=1, ln=True)
    res = pdf.output()
    return bytes(res) if isinstance(res, (bytes, bytearray)) else bytes(res, encoding='latin-1')

# --- EXPORTAÇÃO DOSSIÊ COMPLETO EM EXCEL (4 ABAS DETALHADAS) ---
def gerar_excel_dossie_4abas(lista_empresas):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    title_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    regular_font = Font(name="Calibri", size=11)
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    # ABA 1: RESUMO CADASTRAL
    ws1 = wb.active
    ws1.title = "Resumo Cadastral"
    ws1.merge_cells("A1:O1")
    ws1["A1"] = "DOSSIÊ DE ONBOARDING CONTÁBIL COMPLETO"
    ws1["A1"].font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    ws1["A1"].fill = title_fill
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers1 = ["CNPJ", "Razão Social", "Nome Fantasia", "Situação Cadastral", "Regime", "E-mail", "Telefone", "CNAE Principal", "Anexo Simples", "Capital Social", "Inscrição Municipal", "Inscrição Estadual", "Endereço", "Cód. IBGE", "Link Google Maps"]
    ws1.row_dimensions[3].height = 25
    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=3, column=col_idx, value=h)
        cell.font = white_bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, emp in enumerate(lista_empresas, 4):
        regime_str = "MEI" if emp["opcao_mei"] else ("Simples Nacional" if emp["opcao_simples"] else "Lucro Presumido")
        ies_str = ", ".join(emp["ies"]) if emp["ies"] else "Isento"
        end_str = f"{emp['logradouro']}, {emp['numero']} - {emp['bairro']}, {emp['municipio']}/{emp['uf']}"
        
        vals = [
            emp["cnpj"], emp["razao_social"], emp["nome_fantasia"], emp["situacao"],
            regime_str, emp["email"], emp["telefone"], emp["cnae_principal_str"], emp["diag_principal"]["anexo"],
            emp["capital_social"], emp["im"], ies_str, end_str, emp["cod_ibge"], emp["maps_url"]
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border

    # ABA 2: ENGENHARIA TRIBUTÁRIA E MONOFÁSICOS
    ws2 = wb.create_sheet(title="Análise Tributária & Monofásicos")
    headers2 = ["CNPJ", "Razão Social", "Tipo Atividade", "CNAE", "Descrição", "Anexo do Simples", "Oportunidade Fiscal / Diagnóstico Especializado"]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = white_bold
        cell.fill = header_fill

    curr_row = 2
    for emp in lista_empresas:
        p_cod = emp["cnae_principal_str"].split(" - ")[0]
        p_desc = emp["cnae_principal_str"].split(" - ")[1]
        ws2.cell(row=curr_row, column=1, value=emp["cnpj"]).border = thin_border
        ws2.cell(row=curr_row, column=2, value=emp["razao_social"]).border = thin_border
        ws2.cell(row=curr_row, column=3, value="PRINCIPAL").border = thin_border
        ws2.cell(row=curr_row, column=4, value=p_cod).border = thin_border
        ws2.cell(row=curr_row, column=5, value=p_desc).border = thin_border
        ws2.cell(row=curr_row, column=6, value=emp["diag_principal"]["anexo"]).border = thin_border
        ws2.cell(row=curr_row, column=7, value=emp["diag_principal"]["dica_engenharia"]).border = thin_border
        curr_row += 1
        
        for item in emp["cnaes_secundarios_analise"]:
            ws2.cell(row=curr_row, column=1, value=emp["cnpj"]).border = thin_border
            ws2.cell(row=curr_row, column=2, value=emp["razao_social"]).border = thin_border
            ws2.cell(row=curr_row, column=3, value="SECUNDÁRIO").border = thin_border
            ws2.cell(row=curr_row, column=4, value=item["code"]).border = thin_border
            ws2.cell(row=curr_row, column=5, value=item["desc"]).border = thin_border
            ws2.cell(row=curr_row, column=6, value=item["diag"]["anexo"]).border = thin_border
            ws2.cell(row=curr_row, column=7, value=item["diag"]["dica_engenharia"]).border = thin_border
            curr_row += 1

    # ABA 3: COMPLIANCE E REGULARIDADE CADASTRAL
    ws3 = wb.create_sheet(title="Compliance & Regularidade Cadastral")
    headers3 = ["CNPJ", "Razão Social", "Situação Receita Federal", "Status e-CAC", "Status FGTS", "CNDT Trabalhista", "Apontamentos"]
    for col_idx, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = white_bold
        cell.fill = header_fill

    curr_row = 2
    for emp in lista_empresas:
        c = emp["compliance"]
        ws3.cell(row=curr_row, column=1, value=emp["cnpj"]).border = thin_border
        ws3.cell(row=curr_row, column=2, value=emp["razao_social"]).border = thin_border
        ws3.cell(row=curr_row, column=3, value=c["cnd_federal"]).border = thin_border
        ws3.cell(row=curr_row, column=4, value=c["val_federal"]).border = thin_border
        ws3.cell(row=curr_row, column=5, value=c["cnd_fgts"]).border = thin_border
        ws3.cell(row=curr_row, column=6, value=c["cndt_trabalhista"]).border = thin_border
        ws3.cell(row=curr_row, column=7, value=c["processos_judiciais"]).border = thin_border
        curr_row += 1

    # ABA 4: SÓCIOS E QSA
    ws4 = wb.create_sheet(title="Quadro Societário (QSA)")
    headers4 = ["CNPJ", "Razão Social", "Sócio / Administrador", "Qualificação", "Faixa Etária", "Alerta do Simples Nacional"]
    for col_idx, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col_idx, value=h)
        cell.font = white_bold
        cell.fill = header_fill

    curr_row = 2
    for emp in lista_empresas:
        if emp["qsa"]:
            for s in emp["qsa"]:
                ws4.cell(row=curr_row, column=1, value=emp["cnpj"]).border = thin_border
                ws4.cell(row=curr_row, column=2, value=emp["razao_social"]).border = thin_border
                ws4.cell(row=curr_row, column=3, value=s.get("nome")).border = thin_border
                ws4.cell(row=curr_row, column=4, value=s.get("qualificacao")).border = thin_border
                ws4.cell(row=curr_row, column=5, value=s.get("faixa_etaria")).border = thin_border
                ws4.cell(row=curr_row, column=6, value=s.get("alerta_outras_empresas")).border = thin_border
                curr_row += 1
        else:
            ws4.cell(row=curr_row, column=1, value=emp["cnpj"]).border = thin_border
            ws4.cell(row=curr_row, column=2, value=emp["razao_social"]).border = thin_border
            ws4.cell(row=curr_row, column=3, value="Empresário Individual / MEI").border = thin_border
            ws4.cell(row=curr_row, column=4, value="N/A").border = thin_border
            ws4.cell(row=curr_row, column=5, value="N/A").border = thin_border
            ws4.cell(row=curr_row, column=6, value="Sem sócios").border = thin_border
            curr_row += 1

    wb.save(output)
    output.seek(0)
    return output

# --- EXPORTAÇÃO DOSSIÊ COMPLETO EM PDF ---
def gerar_pdf_dossie_completo(emp):
    pdf = FPDF()
    pdf.add_page()
    
    pdf.set_fill_color(31, 73, 125)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "DOSSIE CONTABIL E DIAGNOSTICO DE COMPLIANCE", ln=True, align="C", fill=True)
    pdf.ln(4)
    
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 6, tratar_texto_pdf(f"Razao Social: {emp['razao_social']}"), ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, tratar_texto_pdf(f"CNPJ: {emp['cnpj']} | Fantasia: {emp['nome_fantasia']}"), ln=True)
    pdf.cell(0, 5, tratar_texto_pdf(f"Contato: Tel {emp['telefone']} | Email: {emp['email']}"), ln=True)
    pdf.cell(0, 5, tratar_texto_pdf(f"Situacao: {emp['situacao']} | Porte: {emp['porte']} | Capital: R$ {emp['capital_social']:,.2f}"), ln=True)
    pdf.cell(0, 5, tratar_texto_pdf(f"Endereco: {emp['logradouro']}, {emp['numero']} - {emp['bairro']}, {emp['municipio']}/{emp['uf']} (IBGE: {emp['cod_ibge']})"), ln=True)
    pdf.ln(3)

    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(0, 6, "1. Regularidade Cadastral na Receita Federal", ln=True, fill=True)
    pdf.set_font("Helvetica", "", 9)
    comp = emp["compliance"]
    pdf.cell(0, 5, tratar_texto_pdf(f"Status Receita Federal: {comp['cnd_federal']} ({comp['val_federal']})"), ln=True)
    pdf.cell(0, 5, tratar_texto_pdf(f"Status FGTS: {comp['cnd_fgts']}"), ln=True)
    pdf.cell(0, 5, tratar_texto_pdf(f"CNDT Trabalhista: {comp['cndt_trabalhista']}"), ln=True)
    pdf.cell(0, 5, tratar_texto_pdf(f"Apontamentos: {comp['processos_judiciais']}"), ln=True)
    pdf.ln(3)

    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "2. Diagnostico Tributario e Oportunidade Fiscal", ln=True, fill=True)
    pdf.set_font("Helvetica", "", 9)
    p_diag = emp["diag_principal"]
    pdf.cell(0, 5, tratar_texto_pdf(f"CNAE Principal: {emp['cnae_principal_str']}"), ln=True)
    pdf.cell(0, 5, tratar_texto_pdf(f"Enquadramento: {p_diag['anexo']} (Aliquota: {p_diag['aliquota_inicial']})"), ln=True)
    pdf.multi_cell(0, 5, tratar_texto_pdf(f"Orientacao do Especialista: {p_diag['dica_engenharia']}"))
    pdf.ln(3)

    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "3. Quadro Societario (QSA) e Alertas", ln=True, fill=True)
    pdf.set_font("Helvetica", "", 9)
    if emp['qsa']:
        for s in emp['qsa']:
            pdf.cell(0, 5, tratar_texto_pdf(f"- {s.get('nome')} ({s.get('qualificacao')})"), ln=True)
    else:
        pdf.cell(0, 5, "Empresario Individual / MEI sem socios.", ln=True)
        
    if emp['tem_risco_societario']:
        pdf.cell(0, 5, "ALERTA: Empresa com multiplos socios. Checar participacao em outras empresas no Simples.", ln=True)
    pdf.ln(3)

    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "4. Checklist de Onboarding Contabil", ln=True, fill=True)
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 5, "[ ] Para Varejo/Mini Mercados: Exigir relatorio NCM para apurar PIS/COFINS Monofasico e ICMS ST.", ln=True)
    pdf.cell(0, 5, "[ ] Para Servicos: Checar folha de pagamento para aplicar Fator R (Atingir 28% para tributar em 6%).", ln=True)
    pdf.cell(0, 5, "[ ] Validar somatorio de faturamento do socio em outras empresas do Simples (Teto R$ 4.8 mi).", ln=True)

    res = pdf.output()
    return bytes(res) if isinstance(res, (bytes, bytearray)) else bytes(res, encoding='latin-1')

# --- GERADOR DE PROPOSTA COMERCIAL EM PDF (MODELO FIEL MERCABILIZA) ---
def gerar_proposta_minimercado_pdf(emp, incluir_desenq, incluir_abertura, num_cnpjs, num_pessoas):
    pdf = FPDF()
    pdf.add_page()
    
    # Cabeçalho Mercabiliza
    pdf.set_fill_color(220, 50, 80) # Tom vermelho/rosa institucional Mercabiliza
    pdf.rect(0, 0, 210, 12, 'F')
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 6, "MERCABILIZA - CONTABILIDADE PARA MINIMERCADOS", ln=True, align="C")
    pdf.ln(8)
    
    # Título da Proposta
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, "PROPOSTA DE PRESTACAO DE SERVICOS CONTABEIS", ln=True, align="L")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, tratar_texto_pdf(f"Cliente: {emp['razao_social']} | CNPJ: {emp['cnpj']}"), ln=True)
    pdf.cell(0, 5, tratar_texto_pdf(f"Cidade/UF: {emp['municipio']}/{emp['uf']} | Contato: {emp['telefone']}"), ln=True)
    pdf.ln(4)

    # Texto Institucional (Fiel aos PDFs originais)
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(0, 6, "Sobre a Mercabiliza", ln=True, fill=True)
    pdf.set_font("Helvetica", "", 8)
    pdf.multi_cell(0, 4, tratar_texto_pdf(
        "Especialista em solucoes contabeis para minimercados autonomos, a Mercabiliza nasceu com um unico proposito: "
        "apoiar operadores com inteligencia contabil, seguranca trabalhista e estrategias que impulsionam o crescimento. "
        "Oferecemos contabilidade com foco em performance, analises tributarias personalizadas, departamento pessoal "
        "preventivo e processos otimizados com tecnologia."
    ))
    pdf.ln(3)

    # Seção 1: Escopo dos Serviços
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "1. Escopo dos Servicos Prestados", ln=True, fill=True)
    pdf.set_font("Helvetica", "", 8)
    
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(0, 5, "1.1 AREA CONTABIL:", ln=True)
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 4, tratar_texto_pdf("- Classificacao, registro e escrituracao contabil de todas as operacoes financeiras e patrimoniais."), ln=True)
    pdf.cell(0, 4, tratar_texto_pdf("- Elaboracao do Balanco Patrimonial, DRE e apuracao dos resultados."), ln=True)
    pdf.cell(0, 4, tratar_texto_pdf("- Entrega das obrigacoes acessorias contabeis (ECD/ECF quando aplicavel)."), ln=True)
    pdf.ln(2)

    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(0, 5, "1.2 AREA FISCAL:", ln=True)
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 4, tratar_texto_pdf("- Escrituracao fiscal completa e apuracao do Simples Nacional com segregacao de PIS/COFINS Monofasico."), ln=True)
    pdf.cell(0, 4, tratar_texto_pdf("- Elaboracao e entrega do SPED, DCTF, EFD-Reinf, GIA, DAS e DASN."), ln=True)
    pdf.cell(0, 4, tratar_texto_pdf("- Atendimento consultivo para planejamento tributario de bebidas e conveniência."), ln=True)
    pdf.ln(2)

    if num_pessoas > 0:
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(0, 5, "1.3 DEPARTAMENTO PESSOAL:", ln=True)
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(0, 4, tratar_texto_pdf("- Gestao de empregados, admissoes, rescisoes e controle de folha em conformidade com a CLT."), ln=True)
        pdf.cell(0, 4, tratar_texto_pdf("- Emissao das guias de encargos sociais (INSS, FGTS, IRRF) e transmissoes do eSocial/DCTFWeb."), ln=True)
        pdf.ln(2)

    # Seção 2: Resumo Financeiro da Proposta
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "2. Investimento e Honorarios", ln=True, fill=True)
    pdf.set_font("Helvetica", "", 9)

    # Cálculos Dinâmicos
    val_base = 350.0
    val_cnpjs_add = max(0, num_cnpjs - 1) * 50.0
    
    # Regra DP: R$ 50 a cada bloco de até 3 pessoas
    blocos_dp = (num_pessoas + 2) // 3 if num_pessoas > 0 else 0
    val_dp = blocos_dp * 50.0
    
    total_mensal = val_base + val_cnpjs_add + val_dp

    pdf.cell(0, 5, tratar_texto_pdf(f"- Honorarios Contabeis Recorrentes (Mensalidade Base): R$ {val_base:,.2f} / mes"), ln=True)
    if val_cnpjs_add > 0:
        pdf.cell(0, 5, tratar_texto_pdf(f"- Adicional por Unidades/CNPJs ({num_cnpjs - 1} filiais x R$ 50): R$ {val_cnpjs_add:,.2f} / mes"), ln=True)
    if val_dp > 0:
        pdf.cell(0, 5, tratar_texto_pdf(f"- Adicional Departamento Pessoal ({num_pessoas} vinculos ativos / {blocos_dp} bloco(s)): R$ {val_dp:,.2f} / mes"), ln=True)

    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 7, tratar_texto_pdf(f"TOTAL MENSALIDADE RECORRENTE: R$ {total_mensal:,.2f} / mes"), ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.ln(2)

    # Serviços Pontuais
    if incluir_desenq or incluir_abertura:
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(0, 5, "Servicos Pontuais / Taxa Unica de Implantacao:", ln=True)
        pdf.set_font("Helvetica", "", 9)
        if incluir_desenq:
            pdf.cell(0, 5, tratar_texto_pdf("- Processo de Desenquadramento de MEI: R$ 350,00 (parcela unica)"), ln=True)
        if incluir_abertura:
            pdf.cell(0, 5, tratar_texto_pdf("- Constituicao / Abertura de Empresa: R$ 1.600,00 (parcela unica)"), ln=True)
        pdf.ln(2)

    pdf.set_font("Helvetica", "I", 8)
    pdf.multi_cell(0, 4, tratar_texto_pdf("Servicos extras como Imposto de Renda Pessoa Fisica, alteracoes contratuais complexas e licencas especificas serao cotados a parte. Apos o aceite, enviaremos o Contrato de Prestacao de Servicos formal."))
    pdf.ln(4)

    # Assinatura
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(0, 4, "Luis Felipe - Socio Mercabiliza Contabilidade", ln=True)
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 4, "Telefone: +55 19 99285-3550 | E-mail: luisfelipe@contabilidadeclassea.com.br", ln=True)
    pdf.cell(0, 4, tratar_texto_pdf(f"Proposta emitida em {datetime.date.today().strftime('%d/%m/%Y')} - Valida por 15 dias."), ln=True)

    res = pdf.output()
    return bytes(res) if isinstance(res, (bytes, bytearray)) else bytes(res, encoding='latin-1')

# --- RENDERIZADOR DOS 4 PAINÉIS ---
def renderizar_paineis_dossie(d):
    c = d["compliance"]
    p_diag = d["diag_principal"]
    
    st.markdown("---")
    st.header(f"🏢 {d['razao_social']}")
    st.caption(f"CNPJ: {d['cnpj']} | Abertura: {d['data_abertura']} | Tipo: {d['matriz_filial']} | Região: {d['regiao_macro']}")

    with st.expander("📜 Ver Comprovante de Inscrição e Situação Cadastral (Cartão CNPJ Oficial)"):
        renderizar_cartao_cnpj_html(d)
        pdf_cartao = gerar_pdf_cartao_cnpj_oficial(d)
        st.download_button(
            label="📄 Baixar Cartão CNPJ Oficial em PDF",
            data=pdf_cartao,
            file_name=f"cartao_cnpj_{d['cnpj']}.pdf",
            mime="application/pdf",
            key=f"btn_cartao_{d['cnpj']}"
        )
    
    # PAINEL 1: REGULARIDADE CADASTRAL E COMPLIANCE
    st.subheader("🛡️ Painel 1: Compliance e Situação Cadastral na Receita")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Situação do CNPJ", d['situacao'])
        st.caption(c["cnd_federal"])
    with col2:
        st.metric("Cadastro FGTS", "Ativo")
        st.caption(c["cnd_fgts"])
    with col3:
        st.metric("CNDT Trabalhista", "Ativo")
        st.caption(c["cndt_trabalhista"])
    with col4:
        st.metric("Apontamentos", "0 Públicos")
        st.caption(c["processos_judiciais"])

    st.info("ℹ️ **Nota de Compliance:** O status cadastral indica se a empresa está ativa e operacional perante a Receita Federal. A varredura de débitos fiscais em aberto (Certidão Negativa de Débitos/e-CAC) exige emissão formal com certificado digital do cliente.")

    st.markdown("---")
    
    # PAINEL 2: ENGENHARIA TRIBUTÁRIA E DIAGNÓSTICO
    st.subheader("⚡ Painel 2: Engenharia Tributária & Oportunidades Fiscais")
    col_fat1, col_fat2 = st.columns([1, 2])
    with col_fat1:
        st.markdown(f"**CNAE Principal:** `{d['cnae_principal_str']}`")
        st.success(f"**Enquadramento:** {p_diag['anexo']}")
        st.write(f"**Alíquota Inicial:** {p_diag['aliquota_inicial']}")
        if p_diag.get("is_minimercado"):
            st.info("🛒 **Empresa do Ramo de Mini Mercado / Varejo Alimentício**")

    with col_fat2:
        st.warning(f"💡 **Diagnóstico Especializado de Economia:**\n\n{p_diag['dica_engenharia']}")

    with st.expander(f"📋 Análise dos {len(d['cnaes_secundarios_analise'])} CNAEs Secundários"):
        if d['cnaes_secundarios_analise']:
            sec_df = pd.DataFrame([{
                "CNAE": item["code"],
                "Descrição": item["desc"],
                "Anexo Estimado": item["diag"]["anexo"],
                "Alíquota": item["diag"]["aliquota_inicial"],
                "Orientações Fiscais": item["diag"]["dica_engenharia"]
            } for item in d['cnaes_secundarios_analise']])
            st.dataframe(sec_df, use_container_width=True)
        else:
            st.write("Sem atividades secundárias registradas.")

    st.markdown("---")

    # PAINEL 3: QUADRO SOCIETÁRIO E RISCOS
    st.subheader("👥 Painel 3: Quadro Societário (QSA) e Risco do Teto do Simples")
    if d['tem_risco_societario']:
        st.warning("⚠️ **ALERTA DE RISCO SOCIETÁRIO:** Empresa com múltiplos sócios. Verifique se os sócios possuem participação (>10%) em outras empresas do Simples Nacional para evitar que o faturamento somado ultrapasse o limite legal de R$ 4,8 milhões/ano!")
    if d['qsa']:
        st.table(pd.DataFrame(d['qsa']))
    else:
        st.info("Empresário Individual / MEI sem sócios cadastrados no QSA.")

    st.markdown("---")

    # PAINEL 4: ENDEREÇO, CONTATOS, WHATSAPP DIRECT
    st.subheader("📍 Painel 4: Endereço Fiscal, Contatos & Abordagem Comercial")
    col_e1, col_e2 = st.columns(2)
    with col_e1:
        st.write(f"**Endereço Registrado:** {d['logradouro']}, {d['numero']} - {d['bairro']}, {d['municipio']}/{d['uf']} - CEP: {d['cep']}")
        st.write(f"**Código IBGE:** `{d['cod_ibge']}` | **Macro-região:** `{d['regiao_macro']}`")
        st.markdown(f"[🗺️ Abrir no Google Maps/Street View]({d['maps_url']})")
    with col_e2:
        st.write(f"**E-mail(s):** {d['email']}")
        st.write(f"**Telefone(s):** {d['telefone']}")
        num_limpo = re.sub(r'\D', '', str(d['telefone']))
        if len(num_limpo) >= 10:
            num_wsp = "55" + num_limpo[:11]
            msg_wsp = f"Olá! Sou da Mercabiliza Contabilidade. Analisei o CNPJ {d['cnpj']} ({d['razao_social']}) e identificamos oportunidades de otimização tributária para o seu negócio. Gostaria de receber nosso diagnóstico gratuito?"
            url_wsp = f"https://wa.me/{num_wsp}?text={urllib.parse.quote(msg_wsp)}"
            st.markdown(f"[📱 **Enviar Mensagem no WhatsApp da Empresa**]({url_wsp})")

# --- DEFINIÇÃO E DECLARAÇÃO DAS ABAS PRINCIPAIS ---
aba1, aba2, aba3, aba4, aba5 = st.tabs([
    "🔍 Dossiê Individual Completo", 
    "⚔️ Comparador de Regimes & Economia Monofásica", 
    "📊 Análise em Lote (Upload Excel)", 
    "🛠️ Transição & Calculadora MEI",
    "🗃️ CRM & Banco de Leads"
])


# ==============================================================================
# INÍCIO - ABA 1: DOSSIÊ INDIVIDUAL COMPLETO
# ==============================================================================
with aba1:
    cnpj_input = st.text_input("Digite o CNPJ do Cliente / Mini Mercado:", placeholder="00.000.000/0001-91")

    if st.button("🔍 Gerar Dossiê Inteligente do Cliente", type="primary"):
        cnpj_limpo = limpar_cnpj(cnpj_input)
        
        if not cnpj_limpo:
            st.error("❌ CNPJ inválido. Digite os 14 números corretamente.")
        else:
            with st.spinner("Consultando simultaneamente BrasilAPI, CNPJ.ws, ReceitaWS e IBGE..."):
                dossie = consultar_dossie_completo(cnpj_limpo)
                
            if dossie:
                st.session_state.historico = [e for e in st.session_state.historico if e["cnpj"] != cnpj_limpo]
                st.session_state.historico.insert(0, dossie)
                st.success("✅ Dossiê gerado com dados de 4 fontes sincronizadas e salvo no CRM!")
            else:
                st.error("❌ CNPJ não localizado nas bases públicas.")

    if st.session_state.historico:
        d = st.session_state.historico[0]
        renderizar_paineis_dossie(d)
        
        st.markdown("---")
        st.subheader("📥 Exportação de Relatórios do Cliente")
        down1, down2 = st.columns(2)
        with down1:
            excel_file = gerar_excel_dossie_4abas(st.session_state.historico)
            st.download_button(
                label="📊 Baixar Dossiê Completo no Excel (4 Abas Detalhadas)",
                data=excel_file,
                file_name=f"dossie_contabil_completo_{d['cnpj']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        with down2:
            pdf_bytes = gerar_pdf_dossie_completo(d)
            st.download_button(
                label="📄 Baixar Dossiê Completo em PDF (.pdf)",
                data=pdf_bytes,
                file_name=f"dossie_contabil_completo_{d['cnpj']}.pdf",
                mime="application/pdf",
                type="primary",
                use_container_width=True
            )

        st.markdown("---")
        st.subheader("📄 Gerador de Proposta Comercial Personalizada (Mercabiliza)")
        st.write("Configure os serviços pontuais e recorrentes para montar a proposta oficial em PDF:")

        # SIMULADOR E SELETOR DINÂMICO DE SERVIÇOS DA PROPOSTA
        col_sec1, col_sec2 = st.columns(2)
        with col_sec1:
            st.markdown("### 🛠️ Serviços Pontuais (Cobrança Única)")
            inc_desenq = st.checkbox("Desenquadramento de MEI (R$ 350,00)", value=d.get("opcao_mei", False))
            inc_abertura = st.checkbox("Constituição / Abertura de Empresa (R$ 1.600,00)", value=False)

        with col_sec2:
            st.markdown("### 🔄 Serviços Recorrentes (Mensalidade)")
            qtd_cnpjs = st.number_input("Quantidade de Unidades / CNPJs (Matriz + Filiais):", min_value=1, max_value=20, value=1, step=1)
            qtd_pessoas = st.number_input("Quantidade de Vínculos / Pessoas (Funcionários e Pró-Labore):", min_value=0, max_value=50, value=1, step=1)

        # CÁLCULO PRÉVIO DA PROPOSTA
        v_base = 350.0
        v_cnpjs_add = max(0, qtd_cnpjs - 1) * 50.0
        blocos_dp_calc = (qtd_pessoas + 2) // 3 if qtd_pessoas > 0 else 0
        v_dp = blocos_dp_calc * 50.0
        total_recorrente = v_base + v_cnpjs_add + v_dp

        st.info(
            f"💰 **RESUMO DOS HONORÁRIOS DA PROPOSTA:**\n\n"
            f"• **Mensalidade Recorrente Calculada:** **R$ {total_recorrente:,.2f} / mês** "
            f"(Base R$ 350,00 + R$ {v_cnpjs_add:,.2f} filiais + R$ {v_dp:,.2f} DP)\n"
            f"• **Serviços Pontuais Selecionados:** "
            f"{'Desenquadramento MEI (R$ 350,00) ' if inc_desenq else ''}"
            f"{'Constituição de Empresa (R$ 1.600,00)' if inc_abertura else ('Nenhum' if not inc_desenq else '')}"
        )

        pdf_proposta = gerar_proposta_minimercado_pdf(d, inc_desenq, inc_abertura, qtd_cnpjs, qtd_pessoas)
        
        st.download_button(
            label="📄 Baixar Proposta Comercial Oficial em PDF (.pdf)",
            data=pdf_proposta,
            file_name=f"Proposta_Servicos_Contabeis_{d['cnpj']}.pdf",
            mime="application/pdf",
            type="primary"
        )
# ==============================================================================
# FIM - ABA 1: DOSSIÊ INDIVIDUAL COMPLETO
# ==============================================================================


# ==============================================================================
# INÍCIO - ABA 2: COMPARADOR DE REGIMES & ECONOMIA MONOFÁSICA
# ==============================================================================
with aba2:
    st.header("⚔️ Comparador de Regimes & Calculadora de Economia Monofásica")
    st.write("Simule o impacto fiscal anual e mensal entre **Simples Nacional (com segregação de Monofásicos)** e **Lucro Presumido** para Mini Mercados Autônomos.")
    
    col_c1, col_c2, col_c3 = st.columns(3)
    with col_c1:
        fat_sim = st.number_input("Faturamento Médio Mensal (R$):", value=35000.0, step=5000.0, key="sim_fat")
    with col_c2:
        tipo_lucro_sel = st.radio("O cliente informou a Margem de Lucro em:", ["Líquido", "Bruto"], horizontal=True)
    with col_c3:
        margem_lucro = st.number_input(f"Margem de Lucro {tipo_lucro_sel} Estimada (%):", value=18.0, step=2.0, key="sim_lucro")

    if tipo_lucro_sel == "Bruto":
        margem_real_calc = margem_lucro * 0.30
        lucro_bruto_rs = fat_sim * (margem_lucro / 100.0)
        lucro_liquido_rs = fat_sim * (margem_real_calc / 100.0)
        
        st.warning(
            f"📊 **ANÁLISE DE MARGEM MERCABILIZA:**\n\n"
            f"• **Margem Bruta (Markup de Vendas):** {margem_lucro:.1f}% (R$ {lucro_bruto_rs:,.2f}/mês sobre as vendas)\n"
            f"• **Margem Líquida Real Estimada (DRE):** **{margem_real_calc:.1f}%** (R$ {lucro_liquido_rs:,.2f}/mês sobram limpos no bolso após pagar CMV, aluguel, energia e taxas do totem).\n\n"
            f"👉 *Ajuste técnico realizado para não inflar o resultado real do mini mercado na apuração fiscal!*"
        )

    imp_simp_anual, imp_pres_anual, melhor_reg, econ_anual = comparar_regimes_simples_presumido(fat_sim, margem_lucro, tipo_lucro_sel)
    
    imp_simp_mensal = imp_simp_anual / 12.0
    imp_pres_mensal = imp_pres_anual / 12.0
    econ_mensal = econ_anual / 12.0

    st.markdown("---")
    st.subheader("📊 Comparativo de Impostos (Visão Mensal & Anual)")
    
    m1, m2, m3 = st.columns(3)
    with m1:
        st.metric("Simples Nacional Otimizado", f"R$ {imp_simp_mensal:,.2f} / mês", delta=f"R$ {imp_simp_anual:,.2f} / ano", delta_color="inverse")
        st.caption("Considerando abate de PIS/COFINS Monofásico")
    with m2:
        st.metric("Lucro Presumido", f"R$ {imp_pres_mensal:,.2f} / mês", delta=f"R$ {imp_pres_anual:,.2f} / ano", delta_color="inverse")
        st.caption("Carga média estimada de 5,9% sobre vendas")
    with m3:
        st.metric("Diferença entre Regimes", f"R$ {econ_mensal:,.2f} / mês", delta=f"R$ {econ_anual:,.2f} / ano")
        st.caption(f"Calculado com base na Margem {tipo_lucro_sel}")

    st.success(f"🏆 **MELHOR REGIME ESTIMADO PARA O MINI MERCADO:** `{melhor_reg.upper()}`")

    st.markdown("---")
    st.subheader("🛒 Simulador de Economia Real com Produtos Monofásicos")
    st.write("Mini mercados vendem itens com imposto recolhido antecipadamente na fábrica (bebidas, refrigerantes, água, energéticos, snacks). Indique a participação estimada desses produtos nas vendas totais:")

    pct_monofasico = st.slider("Porcentagem estimada de vendas em Produtos Monofásicos (%):", min_value=10, max_value=90, value=55, step=5)

    vendas_monofasicas_mes = fat_sim * (pct_monofasico / 100.0)
    economia_das_mes = vendas_monofasicas_mes * 0.0125  
    economia_das_ano = economia_das_mes * 12.0

    col_m1, col_m2 = st.columns(2)
    with col_m1:
        st.metric("Faturamento Monofásico Estimado", f"R$ {vendas_monofasicas_mes:,.2f} / mês", delta=f"R$ {vendas_monofasicas_mes * 12:,.2f} / ano", delta_color="off")
        st.caption(f"Correspondente a {pct_monofasico}% do faturamento bruto")
    with col_m2:
        st.metric("Economia Estimada no DAS (Mercabiliza)", f"R$ {economia_das_mes:,.2f} / mês", delta=f"R$ {economia_das_ano:,.2f} / ano")
        st.caption("Dinheiro recuperado direto na guia mensal pela segregação correta do NCM")

    st.info("💡 **Argumento Comercial de Vendas Mercabiliza:**\n\nMostre ao cliente que a economia de **R$ {:.2f}/mês** na guia do DAS cobre grande parte ou a totalidade dos honorários da contabilidade!".format(economia_das_mes))
# ==============================================================================
# FIM - ABA 2: COMPARADOR DE REGIMES & ECONOMIA MONOFÁSICA
# ==============================================================================


# ==============================================================================
# INÍCIO - ABA 3: ANÁLISE EM LOTE (UPLOAD EXCEL)
# ==============================================================================
with aba3:
    st.header("📊 Análise Contábil em Lote (Upload de Planilha)")
    st.write("Envie uma planilha Excel (`.xlsx`) contendo uma coluna chamada **CNPJ** para analisar dezenas de clientes com o **Dossiê Completo de 4 Painéis**.")
    
    uploaded_file = st.file_uploader("Selecione o arquivo Excel (.xlsx):", type=["xlsx"])
    
    if uploaded_file:
        try:
            df_in = pd.read_excel(uploaded_file)
            col_cnpj = [c for c in df_in.columns if "CNPJ" in str(c).upper()]
            
            if not col_cnpj:
                st.error("❌ A planilha deve conter uma coluna com o nome 'CNPJ'.")
            else:
                cnpjs_lista = df_in[col_cnpj[0]].dropna().astype(str).tolist()
                st.info(f"Encontrados {len(cnpjs_lista)} CNPJs para processamento completo.")
                
                if st.button("🚀 Processar Todos os CNPJs em Lote", type="primary"):
                    prog_bar = st.progress(0)
                    lote_temp = []
                    
                    for idx, c_raw in enumerate(cnpjs_lista):
                        c_limp = limpar_cnpj(c_raw)
                        if c_limp:
                            d_lote = consultar_dossie_completo(c_limp)
                            if d_lote:
                                lote_temp.append(d_lote)
                        prog_bar.progress((idx + 1) / len(cnpjs_lista))
                        
                    st.session_state.lote_processado = lote_temp
                    st.success(f"✅ Processamento de {len(lote_temp)} empresa(s) concluído e gravado no CRM!")

        except Exception as e:
            st.error(f"Erro ao ler a planilha: {e}")

    if st.session_state.lote_processado:
        st.markdown("---")
        st.subheader("🔍 Visualizador Individual de Empresas do Lote")
        
        opcoes_empresas = {f"{e['razao_social']} ({e['cnpj']})": e for e in st.session_state.lote_processado}
        escolha = st.selectbox("Selecione uma empresa do lote para examinar os 4 Painéis na tela:", list(opcoes_empresas.keys()))
        
        emp_selecionada = opcoes_empresas[escolha]
        renderizar_paineis_dossie(emp_selecionada)

        st.markdown("---")
        st.subheader("📥 Exportação Consolidada do Lote Processado")
        
        down_l1, down_l2 = st.columns(2)
        with down_l1:
            excel_lote = gerar_excel_dossie_4abas(st.session_state.lote_processado)
            st.download_button(
                label=f"📊 Baixar Excel do Lote Completo ({len(st.session_state.lote_processado)} empresas em 4 Abas)",
                data=excel_lote,
                file_name=f"dossie_lote_completo_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        with down_l2:
            pdf_emp = gerar_pdf_dossie_completo(emp_selecionada)
            st.download_button(
                label=f"📄 Baixar PDF do Dossiê Selecionado ({emp_selecionada['razao_social'][:20]}...)",
                data=pdf_emp,
                file_name=f"dossie_{emp_selecionada['cnpj']}.pdf",
                mime="application/pdf",
                type="primary",
                use_container_width=True
            )
# ==============================================================================
# FIM - ABA 3: ANÁLISE EM LOTE (UPLOAD EXCEL)
# ==============================================================================


# ==============================================================================
# INÍCIO - ABA 4: TRANSIÇÃO & CÁLCULO RETROATIVO MEI
# ==============================================================================
with aba4:
    st.header("🛠️ Diagnóstico e Simulador de Imposto Retroativo do MEI")
    st.write("Calcule a estimativa de impostos retroativos ($R\$) e guias complementares caso o MEI tenha estourado o limite legal.")
    
    selic_v, ipca_v = obter_indicadores_bacen()
    st.caption(f"🏛️ **Indicadores Macroeconômicos (Banco Central do Brasil em Tempo Real):** Taxa Selic Acumulada: `{selic_v}% a.a.` | IPCA Acumulado: `{ipca_v}% a.a.`")

    col_m1, col_m2 = st.columns(2)
    with col_m1:
        fat_mei = st.number_input("Faturamento Acumulado no Ano pelo MEI (R$):", value=92000.0, step=5000.0)
    with col_m2:
        meses_atv = st.slider("Meses de Atividade no Ano Atual:", min_value=1, max_value=12, value=12)

    diag_mei = calcular_imposto_retroativo_mei(fat_mei, meses_atv)
    
    st.markdown("---")
    st.write(f"**Limite Proporcional do MEI para {meses_atv} mês(es):** R$ {diag_mei['limite_prop']:,.2f}")
    
    if diag_mei["excesso"] > 0:
        st.error(f"🔴 **DESENQUADRAMENTO OBRIGATÓRIO:** Faturamento excedeu o limite em **R$ {diag_mei['excesso']:,.2f}** ({diag_mei['pct_excesso']:.1f}% de excesso).")
        
        m_col1, m_col2, m_col3 = st.columns(3)
        with m_col1:
            st.metric("Imposto Retroativo Estimado", f"R$ {diag_mei['imposto_estimado']:,.2f}")
            st.caption("Apurado no Simples Nacional Anexo I")
        with m_col2:
            st.metric("Encargos de Mora (Selic BACEN)", f"R$ {diag_mei['encargos_estimados']:,.2f}")
            st.caption(f"Com base na Selic de {diag_mei['selic_usada']}% a.a.")
        with m_col3:
            st.metric("Total Estimado com Encargos", f"R$ {diag_mei['imposto_total_com_encargos']:,.2f}")
            st.caption("Guia PGDAS-D estimada")

        st.info(f"💡 **Parecer Técnico da Contabilidade:**\n\n{diag_mei['orientacao']}")
    else:
        st.success("🟢 **MEI REGULAR:** O faturamento está dentro do limite proporcional permitido.")

    st.markdown("---")
    st.caption(
        "⚠️ **Aviso de Isenção e Resguardo Técnico:** "
        "Os valores apresentados nesta calculadora constituem uma **estimativa simulatória gerencial** "
        "com base nas informações declaradas pelo usuário e nas alíquotas vigentes da LC 123/2006. "
        "O valor final exato a ser recolhido em guia DARF/DAS é apurado no sistema oficial PGDAS-D/Receita Federal "
        "após a transmissão das declarações (DASN-SIMEI e PGDAS) pela contabilidade habilitada."
    )
# ==============================================================================
# FIM - ABA 4: TRANSIÇÃO & CÁLCULO RETROATIVO MEI
# ==============================================================================


# ==============================================================================
# INÍCIO - ABA 5: CRM & BANCO DE LEADS (SQLITE)
# ==============================================================================
with aba5:
    st.header("🗃️ CRM Contábil & Histórico de Prospects (SQLite)")
    st.write("Todos os CNPJs pesquisados na plataforma são salvos automaticamente no banco de dados local da sua máquina.")
    
    df_crm = carregar_leads_db()
    
    if not df_crm.empty:
        st.subheader(f"📋 Total de Prospects Registrados: {len(df_crm)}")
        st.dataframe(df_crm, use_container_width=True)
        
        buffer_crm = io.BytesIO()
        with pd.ExcelWriter(buffer_crm, engine='openpyxl') as writer:
            df_crm.to_excel(writer, index=False, sheet_name="Leads_CRM")
        buffer_crm.seek(0)
        
        st.download_button(
            label="📥 Exportar Banco de Leads (CRM) em Excel",
            data=buffer_crm,
            file_name=f"crm_leads_contabeis_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
    else:
        st.info("Nenhum lead ou CNPJ consultado até o momento. Faça pesquisas na Aba 1 ou Aba 3 para alimentar a base de dados.")
# ==============================================================================
# FIM - ABA 5: CRM & BANCO DE LEADS (SQLITE)
# ==============================================================================
